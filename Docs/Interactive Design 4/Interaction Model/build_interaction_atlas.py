"""
Build the Interaction Atlas spreadsheet — a teaching + design tool
for the ID4 Interaction Model.

Sheets:
  1. Atoms        — main sheet, one row per interaction atom
  2. Modes        — reference: 4 modes and what's active/killed
  3. Inputs       — catalog of all inputs with signal type
  4. Outputs      — catalog of all outputs by perceptual domain
  5. Presets      — 13 relationship configuration presets

Modes (Signal × Relationship):
  - Tracking          = Continuous + Bound
  - Threshold Trigger = Continuous + Unbound
  - Switch            = Binary + Bound
  - One-Shot          = Binary + Unbound

Excel features used:
  - Data validation dropdowns (Input, Relationship, Curve, Output, Domain, Feel)
  - Formulas: Signal auto-fills from Input, Mode auto-fills from Signal×Relationship
  - Conditional formatting: killed cells gray out based on Mode
  - Freeze panes, column widths, color coding
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

# ── Palette ──────────────────────────────────────────────────────────

WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
DARK_GRAY   = "D9D9D9"
KILLED_BG   = "E0E0E0"
KILLED_FG   = "AAAAAA"

PATH_A_BG   = "DCEEFB"  # blue tint  — Tracking (Continuous+Bound)
PATH_B_BG   = "D5F5E3"  # green tint — Threshold Trigger (Continuous+Unbound)
PATH_C_BG   = "FEF5E7"  # amber tint — Switch (Binary+Bound)
PATH_D_BG   = "FADBD8"  # red tint   — One-Shot (Binary+Unbound)

HEADER_BG   = "2C3E50"
HEADER_FG   = "FFFFFF"
SECTION_BG  = "34495E"

ACCENT_BLUE   = "2980B9"
ACCENT_GREEN  = "27AE60"
ACCENT_AMBER  = "F39C12"
ACCENT_RED    = "E74C3C"

# ── Styles ───────────────────────────────────────────────────────────

header_font  = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
header_fill  = PatternFill("solid", fgColor=HEADER_BG)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

section_font = Font(name="Calibri", bold=True, size=10, color=HEADER_FG)
section_fill = PatternFill("solid", fgColor=SECTION_BG)

body_font    = Font(name="Calibri", size=10)
body_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align   = Alignment(horizontal="left", vertical="center", wrap_text=True)

killed_fill  = PatternFill("solid", fgColor=KILLED_BG)
killed_font  = Font(name="Calibri", size=10, color=KILLED_FG)

thin_border  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

path_fills = {
    "Tracking":          PatternFill("solid", fgColor=PATH_A_BG),
    "Threshold Trigger": PatternFill("solid", fgColor=PATH_B_BG),
    "Switch":            PatternFill("solid", fgColor=PATH_C_BG),
    "One-Shot":          PatternFill("solid", fgColor=PATH_D_BG),
}


def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_body_cell(cell, align=None):
    cell.font = body_font
    cell.alignment = align or body_align
    cell.border = thin_border


# ── Data ─────────────────────────────────────────────────────────────

# Inputs with their natural signal type
INPUTS = [
    ("Proximity",       "Continuous", "Distance to target (m)",          "ProximitySensor"),
    ("Gaze Direction",  "Continuous", "Dot product (0–1)",               "GazeSensor"),
    ("Gaze Duration",   "Continuous", "Accumulated seconds looking",     "GazeSensor"),
    ("Stillness",       "Continuous", "Seconds since last movement",     "StillnessSensor"),
    ("Velocity",        "Continuous", "Player speed (m/s)",              "VelocitySensor"),
    ("Slider",          "Continuous", "UI value (0–1)",                  "World-space UI"),
    ("Time",            "Continuous", "Clock / oscillating",             "Autonomous"),
    ("Random",          "Continuous", "Random float per interval",       "Autonomous"),
    ("Contact",         "Binary",    "Enter/Exit trigger zone",         "TriggerSensor"),
    ("Button",          "Binary",    "Press/Release",                   "World-space UI"),
    ("Toggle",          "Binary",    "On/Off switch",                   "World-space UI"),
    ("Grab",            "Binary",    "Pick up / release object",        "InteractionSystem"),
    ("Throw",           "Binary",    "Release with velocity",           "InteractionSystem"),
]

# Outputs by domain
OUTPUTS = [
    # domain, output name, value type, Unity property
    ("Light",       "Point/Spot Intensity",   "float",  "Light.intensity"),
    ("Light",       "Light Color",            "color",  "Light.color"),
    ("Light",       "Shadow Strength",        "float",  "Light.shadowStrength"),
    ("Light",       "Emission Intensity",     "float",  "Material._EmissionColor intensity"),
    ("Light",       "Emission Color",         "color",  "Material._EmissionColor"),
    ("Material",    "Base Color",             "color",  "Material._BaseColor"),
    ("Material",    "Smoothness",             "float",  "Material._Smoothness"),
    ("Material",    "Metallic",               "float",  "Material._Metallic"),
    ("Material",    "Alpha / Opacity",        "float",  "Material._BaseColor.a"),
    ("Material",    "Dissolve",               "float",  "Material._Dissolve"),
    ("Material",    "Fresnel / Rim Glow",     "float",  "Material._FresnelPower"),
    ("Material",    "Normal Map Strength",    "float",  "Material._BumpScale"),
    ("Material",    "Tiling / UV Offset",     "float",  "Material._MainTex_ST"),
    ("Transform",   "Scale (uniform)",        "float",  "Transform.localScale"),
    ("Transform",   "Scale (per-axis)",       "vec3",   "Transform.localScale.xyz"),
    ("Transform",   "Position",               "vec3",   "Transform.localPosition"),
    ("Transform",   "Rotation",               "vec3",   "Transform.localEulerAngles"),
    ("Transform",   "Vertex Displacement",    "float",  "Shader vertex offset"),
    ("Spawn",       "Object Activation",      "bool",   "GameObject.SetActive"),
    ("Spawn",       "Spawn Object",           "event",  "Instantiate"),
    ("Spawn",       "Destroy Object",         "event",  "Destroy"),
    ("Particles",   "Emission Rate",          "float",  "ParticleSystem.emission.rate"),
    ("Particles",   "Particle Size",          "float",  "ParticleSystem.main.startSize"),
    ("Particles",   "Particle Speed",         "float",  "ParticleSystem.main.startSpeed"),
    ("Particles",   "Particle Color",         "color",  "ParticleSystem.main.startColor"),
    ("Particles",   "Burst",                  "event",  "ParticleSystem.Emit()"),
    ("Sound",       "Volume",                 "float",  "AudioSource.volume"),
    ("Sound",       "Pitch",                  "float",  "AudioSource.pitch"),
    ("Sound",       "Spatial Blend",          "float",  "AudioSource.spatialBlend"),
    ("Sound",       "Low-Pass Filter",        "float",  "AudioLowPassFilter.cutoff"),
    ("Sound",       "Reverb",                 "float",  "AudioReverbFilter.level"),
    ("Sound",       "Play / Stop",            "bool",   "AudioSource.Play/Stop"),
    ("Environment", "Fog Density",            "float",  "RenderSettings.fogDensity"),
    ("Environment", "Fog Color",              "color",  "RenderSettings.fogColor"),
    ("Environment", "Ambient Intensity",      "float",  "RenderSettings.ambientIntensity"),
    ("Environment", "Ambient Color",          "color",  "RenderSettings.ambientLight"),
    ("Environment", "Skybox Exposure",        "float",  "Skybox._Exposure"),
    ("Post-Proc",   "Bloom Intensity",        "float",  "Volume: Bloom.intensity"),
    ("Post-Proc",   "Color Temperature",      "float",  "Volume: WhiteBalance.temperature"),
    ("Post-Proc",   "Saturation",             "float",  "Volume: ColorAdjustments.saturation"),
    ("Post-Proc",   "Vignette",               "float",  "Volume: Vignette.intensity"),
    ("Post-Proc",   "Depth of Field",         "float",  "Volume: DepthOfField.focusDist"),
    ("Post-Proc",   "Chromatic Aberration",   "float",  "Volume: ChromaticAberration.intensity"),
    ("Post-Proc",   "Film Grain",             "float",  "Volume: FilmGrain.intensity"),
    ("Camera",      "Field of View",          "float",  "Camera.fieldOfView"),
    ("Camera",      "Camera Shake",           "float",  "CinemachineImpulse"),
    ("Camera",      "Camera Position",        "vec3",   "CinemachineVirtualCamera"),
]

# Curve presets per mode
CURVES_BOUND_CONT     = "Linear,Ease-in,Ease-out,Ease-in-out,Inverse,S-curve,Custom"
CURVES_UNBOUND_CONT   = "Swell,Spike,Ramp-up,Ramp-down,Custom"
CURVES_BOUND_BIN      = "—"   # no curve for binary bound
CURVES_UNBOUND_BIN    = "Burst,Swell,Snap+tail,Overshoot,Elastic,Custom"

# Relationship presets (from primitives.md)
PRESETS = [
    # name, signal, relationship, curve, attack, release, duration, description
    ("Linear Tracking",      "Cont", "Bound",   "Linear",    "Instant",  "Instant",  "—", "Output mirrors input proportionally"),
    ("Smoothed Tracking",    "Cont", "Bound",   "Linear",    "Medium",   "Medium",   "—", "Output follows with uniform smoothing"),
    ("Asymmetric Tracking",  "Cont", "Bound",   "Linear",    "Fast",     "Slow",     "—", "Quick response, slow fade — lingering"),
    ("Eased Tracking",       "Cont", "Bound",   "Ease-out",  "Fast",     "Medium",   "—", "Responsive onset, natural settling"),
    ("Inverted Tracking",    "Cont", "Bound",   "Inverse",   "Medium",   "Medium",   "—", "Output opposes input (Rain Room)"),
    ("Binary Snap",          "Bin",  "Bound",   "—",         "Instant",  "Instant",  "—", "Hard switch, mechanical feel"),
    ("Binary Fade",          "Bin",  "Bound",   "—",         "Medium",   "Medium",   "—", "Smooth transition between states"),
    ("Binary Asymmetric",    "Bin",  "Bound",   "—",         "Fast",     "Slow",     "—", "Quick on, lingering off"),
    ("Burst",                "Bin",  "Unbound", "Burst",     "—",        "—",        "0.2s", "Instant peak, fast decay — impact"),
    ("Swell",                "Bin",  "Unbound", "Swell",     "—",        "—",        "1.5s", "Gradual rise and fall — organic"),
    ("Snap and Tail",        "Bin",  "Unbound", "Snap+tail", "—",        "—",        "1.0s", "Instant peak, slow lingering decay"),
    ("Overshoot",            "Bin",  "Unbound", "Overshoot", "—",        "—",        "0.8s", "Peak overshoots then settles — bouncy"),
    ("Timed Playback",       "Bin",  "Unbound", "Custom",    "—",        "—",        "var", "Arbitrary curve over time"),
]

# Example atoms for the Atoms sheet
EXAMPLE_ATOMS = [
    # name, input, signal, rel, curve, in_min, in_max, out_min, out_max, attack, release, duration, output, domain, feel
    # --- TRACKING: Continuous + Bound ---
    ("Proximity Glow",            "Proximity", "Continuous", "Bound",   "Linear",   "0",  "5",  "0", "1",  "—",      "—",      "—",    "Emission Intensity",  "Light",       "Attentive, crisp"),
    ("Proximity Glow (eased)",    "Proximity", "Continuous", "Bound",   "Ease-out", "0",  "5",  "0", "1",  "—",      "—",      "—",    "Emission Intensity",  "Light",       "Attentive, STICKY"),
    ("Inverse Glow (Rain Room)",  "Proximity", "Continuous", "Bound",   "Inverse",  "0",  "5",  "1", "0",  "—",      "—",      "—",    "Emission Intensity",  "Light",       "RELUCTANT, evasive"),
    ("Smoothed Proximity Glow",   "Proximity", "Continuous", "Bound",   "Linear",   "0",  "5",  "0", "1",  "Fast",   "Slow",   "—",    "Emission Intensity",  "Light",       "Eager in, LINGERING out"),
    ("Gaze Reveal",               "Gaze Direction", "Continuous", "Bound", "Ease-in", "0.5","1",  "0", "1", "Slow",   "Fast",   "—",    "Alpha / Opacity",     "Material",    "Reluctant to show, crisp to hide"),
    ("Stillness Bloom",           "Stillness", "Continuous", "Bound",   "Ease-in",  "0",  "5",  "0", "1",  "Slow",   "Fast",   "—",    "Bloom Intensity",     "Post-Proc",   "Rewards patience"),
    # --- SWITCH: Binary + Bound ---
    ("Zone Light ON / OFF (snap)",     "Contact",   "Binary",     "Bound",   "—",        "—",  "—",  "0", "1",  "Instant","Instant", "—",   "Point/Spot Intensity","Light",       "Mechanical, crisp"),
    ("Zone Light ON / OFF (linger)",   "Contact",   "Binary",     "Bound",   "—",        "—",  "—",  "0", "1",  "Fast",   "Slow",    "—",   "Point/Spot Intensity","Light",       "LINGERING, memory"),
    ("Zone Light ON / OFF (reluctant)","Contact",   "Binary",     "Bound",   "—",        "—",  "—",  "0", "1",  "Slow",   "Fast",    "—",   "Point/Spot Intensity","Light",       "RELUCTANT, suspicious"),
    ("Presence Fog",              "Contact",   "Binary",     "Bound",   "—",        "—",  "—",  "0", "0.05","Slow",  "Slow",    "—",   "Fog Density",         "Environment", "Atmospheric, enveloping"),
    # --- ONE-SHOT: Binary + Unbound ---
    ("Button Play Animation",     "Button",    "Binary",     "Unbound", "Burst",    "—",  "—",  "0", "1",  "—",      "—",       "0.3s", "Object Activation",  "Spawn",       "Explosive, eager"),
    ("Button Swell",              "Button",    "Binary",     "Unbound", "Swell",    "—",  "—",  "0", "1",  "—",      "—",       "1.5s", "Emission Intensity", "Light",       "Organic, breathing"),
    ("Contact Particle Burst",    "Contact",   "Binary",     "Unbound", "Overshoot","—",  "—",  "0", "1",  "—",      "—",       "0.8s", "Burst",             "Particles",   "Playful, bouncy"),
    ("Grab Flash",                "Grab",      "Binary",     "Unbound", "Snap+tail","—",  "—",  "0", "1",  "—",      "—",       "1.0s", "Emission Intensity", "Light",       "Affirming, then fading"),
    # --- THRESHOLD TRIGGER: Continuous + Unbound ---
    ("Gaze Dissolve",             "Gaze Duration", "Continuous", "Unbound", "Swell",  "—",  "3",  "0", "1", "—",      "—",       "2.0s", "Dissolve",           "Material",    "Atmospheric, delayed"),
    ("Gaze Camera Shake",         "Gaze Duration", "Continuous", "Unbound", "Burst",  "—",  "3",  "0", "1", "—",      "—",       "0.5s", "Camera Shake",       "Camera",      "Startling, impact"),
    # --- EMPTY ROWS for students ---
]

# Coupling quality vocabulary
FEEL_OPTIONS = "Eager,Reluctant,Attentive,Indifferent,Precise,Forgiving,Sticky,Crisp,Alive,Still"


# ═════════════════════════════════════════════════════════════════════
#  BUILD WORKBOOK
# ═════════════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()

# ── SHEET 1: ATOMS (main) ───────────────────────────────────────────

ws = wb.active
ws.title = "Atoms"
ws.sheet_properties.tabColor = "2980B9"

# Column config: (header, width, sub_header_or_note)
ATOM_COLS = [
    ("#",              4,   ""),
    ("Name",          22,   "descriptive label"),
    ("Input",         16,   "what triggers it"),
    ("Signal",        12,   "auto from Input"),
    ("Relationship",  14,   "Bound / Unbound"),
    ("Mode",          18,   "auto from Sig×Rel"),
    ("Curve",         14,   "mapping shape"),
    ("In Min",        10,   "range start"),
    ("In Max /\nThreshold", 12, "range end or threshold"),
    ("Out Min /\nOFF",10,   "minimum or OFF value"),
    ("Out Max /\nON / Peak", 12, "maximum, ON, or peak"),
    ("Attack",        10,   "onset speed"),
    ("Release",       10,   "fade speed"),
    ("Duration",      10,   "playback length"),
    ("Output",        20,   "what changes"),
    ("Domain",        12,   "perceptual category"),
    ("Coupling Feel", 24,   "how does it feel?"),
]

NUM_COLS = len(ATOM_COLS)
HEADER_ROW = 1
SUBHEADER_ROW = 2
DATA_START = 3
NUM_DATA_ROWS = 40  # room for examples + student rows

# Headers
for c, (name, width, note) in enumerate(ATOM_COLS, 1):
    cell = ws.cell(row=HEADER_ROW, column=c, value=name)
    ws.column_dimensions[get_column_letter(c)].width = width

style_header(ws, HEADER_ROW, NUM_COLS)

# Sub-headers (notes row)
for c, (_, _, note) in enumerate(ATOM_COLS, 1):
    cell = ws.cell(row=SUBHEADER_ROW, column=c, value=note)
    cell.font = Font(name="Calibri", size=8, italic=True, color="888888")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    cell.border = thin_border

ws.freeze_panes = "A3"

# ── Formulas for each data row ──

# Build input-signal lookup
continuous_inputs = [i[0] for i in INPUTS if i[1] == "Continuous"]
binary_inputs     = [i[0] for i in INPUTS if i[1] == "Binary"]

for r in range(DATA_START, DATA_START + NUM_DATA_ROWS):
    row_num = r - DATA_START + 1

    # Col A: row number
    ws.cell(row=r, column=1, value=row_num)
    style_body_cell(ws.cell(row=r, column=1))

    # Col D (Signal): auto-fill from Input (col C)
    # =IF(C3="","", IF(OR(C3="Proximity",C3="Gaze Direction",...), "Continuous", "Binary"))
    cont_checks = ",".join([f'C{r}="{inp}"' for inp in continuous_inputs])
    formula_signal = f'=IF(C{r}="","",IF(OR({cont_checks}),"Continuous","Binary"))'
    ws.cell(row=r, column=4, value=formula_signal)

    # Col F (Mode): auto-fill from Signal×Relationship
    formula_path = (
        f'=IF(OR(D{r}="",E{r}=""),"",IF(AND(D{r}="Continuous",E{r}="Bound"),"Tracking",'
        f'IF(AND(D{r}="Continuous",E{r}="Unbound"),"Threshold Trigger",'
        f'IF(AND(D{r}="Binary",E{r}="Bound"),"Switch",'
        f'IF(AND(D{r}="Binary",E{r}="Unbound"),"One-Shot","")))))'
    )
    ws.cell(row=r, column=6, value=formula_path)

    # Style all cells in row
    for c in range(1, NUM_COLS + 1):
        cell = ws.cell(row=r, column=c)
        style_body_cell(cell, left_align if c in (2, 17) else body_align)

# ── Data Validation (dropdowns) ──

# Input dropdown
dv_input = DataValidation(
    type="list",
    formula1='"' + ",".join([i[0] for i in INPUTS]) + '"',
    allow_blank=True,
)
dv_input.prompt = "Select an input source"
dv_input.promptTitle = "Input"
ws.add_data_validation(dv_input)
dv_input.add(f"C{DATA_START}:C{DATA_START + NUM_DATA_ROWS - 1}")

# Relationship dropdown
dv_rel = DataValidation(
    type="list",
    formula1='"Bound,Unbound"',
    allow_blank=True,
)
dv_rel.prompt = "Bound = output follows input. Unbound = output plays independently."
dv_rel.promptTitle = "Relationship"
ws.add_data_validation(dv_rel)
dv_rel.add(f"E{DATA_START}:E{DATA_START + NUM_DATA_ROWS - 1}")

# Curve dropdown (all options — path-specific filtering would need VBA, so we list all)
all_curves = "Linear,Ease-in,Ease-out,Ease-in-out,Inverse,S-curve,Burst,Swell,Spike,Snap+tail,Overshoot,Elastic,Ramp-up,Ramp-down,Custom,—"
dv_curve = DataValidation(
    type="list",
    formula1=f'"{all_curves}"',
    allow_blank=True,
)
dv_curve.prompt = "Tracking: mapping curves. Threshold Trigger/One-Shot: response shapes. Switch: use '—'"
dv_curve.promptTitle = "Curve"
ws.add_data_validation(dv_curve)
dv_curve.add(f"G{DATA_START}:G{DATA_START + NUM_DATA_ROWS - 1}")

# Attack / Release dropdown
speed_options = "Instant,Fast,Medium,Slow,Very Slow,—"
dv_attack = DataValidation(type="list", formula1=f'"{speed_options}"', allow_blank=True)
dv_attack.prompt = "How fast does the output reach its target? (Bound paths only)"
dv_attack.promptTitle = "Attack Speed"
ws.add_data_validation(dv_attack)
dv_attack.add(f"L{DATA_START}:L{DATA_START + NUM_DATA_ROWS - 1}")

dv_release = DataValidation(type="list", formula1=f'"{speed_options}"', allow_blank=True)
dv_release.prompt = "How fast does the output return to rest? (Bound paths only)"
dv_release.promptTitle = "Release Speed"
ws.add_data_validation(dv_release)
dv_release.add(f"M{DATA_START}:M{DATA_START + NUM_DATA_ROWS - 1}")

# Output dropdown
output_names = ",".join(sorted(set(o[1] for o in OUTPUTS)))
# Excel data validation has 255 char limit — split if needed
if len(output_names) <= 255:
    dv_output = DataValidation(type="list", formula1=f'"{output_names}"', allow_blank=True)
else:
    # Reference the Outputs sheet instead
    dv_output = DataValidation(type="list", formula1="Outputs!$B$2:$B$50", allow_blank=True)
dv_output.promptTitle = "Output"
ws.add_data_validation(dv_output)
dv_output.add(f"O{DATA_START}:O{DATA_START + NUM_DATA_ROWS - 1}")

# Domain dropdown
domains = "Light,Material,Transform,Spawn,Particles,Sound,Environment,Post-Proc,Camera"
dv_domain = DataValidation(type="list", formula1=f'"{domains}"', allow_blank=True)
ws.add_data_validation(dv_domain)
dv_domain.add(f"P{DATA_START}:P{DATA_START + NUM_DATA_ROWS - 1}")

# ── Conditional Formatting: gray out killed cells ──

# We use formulas referencing the Mode column (F)
# killed_fill is applied when the mode makes a column irrelevant

def add_killed_rule(ws, col_letter, paths_that_kill, start, count):
    """Gray out cells in col_letter when Mode (col F) is one of paths_that_kill."""
    cell_range = f"{col_letter}{start}:{col_letter}{start + count - 1}"
    for path in paths_that_kill:
        rule = FormulaRule(
            formula=[f'$F{start}="{path}"'],
            fill=killed_fill,
            font=killed_font,
        )
        ws.conditional_formatting.add(cell_range, rule)

# Curve (G): killed for Switch (Binary+Bound — no mapping curve)
add_killed_rule(ws, "G", ["Switch"], DATA_START, NUM_DATA_ROWS)

# Input Min (H): only active for Tracking
add_killed_rule(ws, "H", ["Threshold Trigger", "Switch", "One-Shot"], DATA_START, NUM_DATA_ROWS)

# Input Max / Threshold (I): killed for Switch and One-Shot
add_killed_rule(ws, "I", ["Switch", "One-Shot"], DATA_START, NUM_DATA_ROWS)

# Attack (L): killed for Threshold Trigger and One-Shot (encoded in curve for unbound)
add_killed_rule(ws, "L", ["Threshold Trigger", "One-Shot"], DATA_START, NUM_DATA_ROWS)

# Release (M): killed for Threshold Trigger and One-Shot
add_killed_rule(ws, "M", ["Threshold Trigger", "One-Shot"], DATA_START, NUM_DATA_ROWS)

# Duration (N): killed for Tracking and Switch (continuous — no fixed duration)
add_killed_rule(ws, "N", ["Tracking", "Switch"], DATA_START, NUM_DATA_ROWS)

# ── Row tinting by Mode ──

for mode_name, fill in path_fills.items():
    # Tint column B (Name) based on mode — gives visual row identity
    cell_range = f"B{DATA_START}:B{DATA_START + NUM_DATA_ROWS - 1}"
    rule = FormulaRule(formula=[f'$F{DATA_START}="{mode_name}"'], fill=fill)
    ws.conditional_formatting.add(cell_range, rule)

    # Also tint column F (Mode) itself
    cell_range_f = f"F{DATA_START}:F{DATA_START + NUM_DATA_ROWS - 1}"
    rule_f = FormulaRule(formula=[f'$F{DATA_START}="{mode_name}"'], fill=fill)
    ws.conditional_formatting.add(cell_range_f, rule_f)

# ── Fill example data ──

for i, atom in enumerate(EXAMPLE_ATOMS):
    r = DATA_START + i
    # name, input, signal(skip-formula), rel, curve, in_min, in_max, out_min, out_max, attack, release, duration, output, domain, feel
    ws.cell(row=r, column=2,  value=atom[0])   # Name
    ws.cell(row=r, column=3,  value=atom[1])   # Input
    # col 4 (Signal) is formula — skip
    ws.cell(row=r, column=5,  value=atom[3])   # Relationship
    # col 6 (Path) is formula — skip
    ws.cell(row=r, column=7,  value=atom[4])   # Curve
    ws.cell(row=r, column=8,  value=atom[5])   # In Min
    ws.cell(row=r, column=9,  value=atom[6])   # In Max / Threshold
    ws.cell(row=r, column=10, value=atom[7])   # Out Min
    ws.cell(row=r, column=11, value=atom[8])   # Out Max
    ws.cell(row=r, column=12, value=atom[9])   # Attack
    ws.cell(row=r, column=13, value=atom[10])  # Release
    ws.cell(row=r, column=14, value=atom[11])  # Duration
    ws.cell(row=r, column=15, value=atom[12])  # Output
    ws.cell(row=r, column=16, value=atom[13])  # Domain
    ws.cell(row=r, column=17, value=atom[14])  # Coupling Feel


# ── SHEET 2: MODES (reference) ──────────────────────────────────────

ws2 = wb.create_sheet("Modes")
ws2.sheet_properties.tabColor = "8E44AD"

# Title
ws2.merge_cells("A1:G1")
cell = ws2.cell(row=1, column=1, value="THE 4 MODES — Which parameters are active?")
cell.font = Font(name="Calibri", bold=True, size=14, color=HEADER_FG)
cell.fill = header_fill
cell.alignment = Alignment(horizontal="center", vertical="center")

# Matrix header
ws2.cell(row=3, column=1, value="")
ws2.cell(row=3, column=2, value="")
ws2.merge_cells("C3:D3")
ws2.cell(row=3, column=3, value="BOUND")
ws2.cell(row=3, column=3).font = Font(bold=True, size=12)
ws2.cell(row=3, column=3).alignment = Alignment(horizontal="center")
ws2.merge_cells("F3:G3")
ws2.cell(row=3, column=6, value="UNBOUND")
ws2.cell(row=3, column=6).font = Font(bold=True, size=12)
ws2.cell(row=3, column=6).alignment = Alignment(horizontal="center")

# Row labels + content
matrix_data = [
    (4, "CONTINUOUS", "Tracking", PATH_A_BG,
     "Output follows input continuously.\nProximity → Light intensity",
     "Threshold Trigger", PATH_B_BG,
     "Continuous signal crosses threshold,\ntriggers independent playback.\nGaze duration hits 3s → material dissolves"),
    (6, "BINARY", "Switch", PATH_C_BG,
     "Output transitions between ON / OFF.\nEnter zone → Light ON / OFF",
     "One-Shot", PATH_D_BG,
     "Binary event triggers independent\nplayback. Button → Play animation"),
]

for row_idx, signal_label, path_a_name, col_a, desc_a, path_b_name, col_b, desc_b in matrix_data:
    ws2.cell(row=row_idx, column=1, value=signal_label)
    ws2.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
    ws2.cell(row=row_idx, column=1).alignment = Alignment(vertical="top")

    # Bound cell
    ws2.merge_cells(f"C{row_idx}:D{row_idx}")
    cell_b = ws2.cell(row=row_idx, column=3, value=f"{path_a_name}\n\n{desc_a}")
    cell_b.fill = PatternFill("solid", fgColor=col_a)
    cell_b.alignment = Alignment(wrap_text=True, vertical="top")
    cell_b.border = thin_border
    ws2.row_dimensions[row_idx].height = 70

    # Unbound cell
    ws2.merge_cells(f"F{row_idx}:G{row_idx}")
    cell_u = ws2.cell(row=row_idx, column=6, value=f"{path_b_name}\n\n{desc_b}")
    cell_u.fill = PatternFill("solid", fgColor=col_b)
    cell_u.alignment = Alignment(wrap_text=True, vertical="top")
    cell_u.border = thin_border

# Parameter activation table
param_start = 9
ws2.merge_cells(f"A{param_start}:G{param_start}")
cell = ws2.cell(row=param_start, column=1, value="PARAMETER ACTIVATION BY MODE")
cell.font = Font(name="Calibri", bold=True, size=12, color=HEADER_FG)
cell.fill = header_fill
cell.alignment = Alignment(horizontal="center")

param_headers = ["Parameter", "Tracking\n(Cont+Bound)", "Threshold Trigger\n(Cont+Unbound)", "Switch\n(Bin+Bound)", "One-Shot\n(Bin+Unbound)", "Notes"]
ph_row = param_start + 1
for c, h in enumerate(param_headers, 1):
    cell = ws2.cell(row=ph_row, column=c, value=h)
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill("solid", fgColor=DARK_GRAY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 4
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 18
ws2.column_dimensions["E"].width = 4
ws2.column_dimensions["F"].width = 18
ws2.column_dimensions["G"].width = 18

param_data = [
    # parameter,   A,              B,                  C,              D,                  notes
    ("Curve",       "✓ Input→Out",  "✓ Time→Out",       "— killed",     "✓ Time→Out",       "Same tool (AnimationCurve), different X axis"),
    ("Input Min",   "✓ range start","— killed",         "— killed",     "— killed",         "Only meaningful for continuous bound"),
    ("Input Max",   "✓ range end",  "→ Threshold",      "— killed",     "— killed",         "Becomes single threshold for Threshold Trigger"),
    ("Output Min",  "✓ dim value",  "✓ start value",    "= OFF value",  "= rest value",     "Label changes; always present"),
    ("Output Max",  "✓ bright val", "✓ peak value",     "= ON value",   "= peak value",     "Label changes; always present"),
    ("Attack",      "✓ smoothing↑", "— in curve",       "✓ transition→ON","— in curve",     "Bound: separate param. Unbound: baked into curve"),
    ("Decay",       "○ optional",   "○ optional",       "○ optional",   "— in curve",       "Overshoot/settling — advanced, usually skip"),
    ("Sustain",     "auto: tracks", "— killed",         "auto: holds ON","— killed",        "Automatic behavior, not a parameter"),
    ("Release",     "✓ smoothing↓", "— in curve",       "✓ transition→OFF","— killed",      "Bound: separate param. Unbound: baked into curve"),
    ("Duration",    "— killed",     "✓ play length",    "— killed",     "✓ play length",    "Only unbound paths have fixed duration"),
]

for i, (param, a, b, c_val, d, notes) in enumerate(param_data):
    r = ph_row + 1 + i
    for col_idx, val in enumerate([param, a, b, c_val, d, notes], 1):
        cell = ws2.cell(row=r, column=col_idx, value=val)
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left",
                                   vertical="center", wrap_text=True)
        cell.border = thin_border
        # Color killed cells
        if "killed" in str(val):
            cell.fill = killed_fill
            cell.font = killed_font
        elif "optional" in str(val):
            cell.fill = PatternFill("solid", fgColor="FFF9E6")

ws2.column_dimensions["A"].width = 16
for col in ["B", "C", "D", "E", "F"]:
    ws2.column_dimensions[col].width = 18


# ── SHEET 3: INPUTS ─────────────────────────────────────────────────

ws3 = wb.create_sheet("Inputs")
ws3.sheet_properties.tabColor = "27AE60"

input_headers = ["Input", "Signal Type", "Description", "Unity Component", "Typical Range"]
for c, h in enumerate(input_headers, 1):
    ws3.cell(row=1, column=c, value=h)
style_header(ws3, 1, len(input_headers))

ws3.column_dimensions["A"].width = 18
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 35
ws3.column_dimensions["D"].width = 22
ws3.column_dimensions["E"].width = 16

typical_ranges = {
    "Proximity": "0–10 m",
    "Gaze Direction": "0–1 (dot)",
    "Gaze Duration": "0–∞ sec",
    "Stillness": "0–∞ sec",
    "Velocity": "0–10 m/s",
    "Slider": "0–1",
    "Time": "0–∞ sec",
    "Random": "0–1",
    "Contact": "true / false",
    "Button": "press / release",
    "Toggle": "on / off",
    "Grab": "holding / empty",
    "Throw": "release event",
}

for i, (name, sig, desc, component) in enumerate(INPUTS):
    r = i + 2
    ws3.cell(row=r, column=1, value=name)
    ws3.cell(row=r, column=2, value=sig)
    ws3.cell(row=r, column=3, value=desc)
    ws3.cell(row=r, column=4, value=component)
    ws3.cell(row=r, column=5, value=typical_ranges.get(name, ""))
    for c in range(1, 6):
        cell = ws3.cell(row=r, column=c)
        style_body_cell(cell, left_align if c in (3, 4) else body_align)
        # Tint by signal type
        if sig == "Continuous":
            cell.fill = PatternFill("solid", fgColor="EBF5FB")
        else:
            cell.fill = PatternFill("solid", fgColor="FEF9E7")

ws3.freeze_panes = "A2"


# ── SHEET 4: OUTPUTS ────────────────────────────────────────────────

ws4 = wb.create_sheet("Outputs")
ws4.sheet_properties.tabColor = "E74C3C"

output_headers = ["Domain", "Output", "Value Type", "Unity Property"]
for c, h in enumerate(output_headers, 1):
    ws4.cell(row=1, column=c, value=h)
style_header(ws4, 1, len(output_headers))

ws4.column_dimensions["A"].width = 14
ws4.column_dimensions["B"].width = 24
ws4.column_dimensions["C"].width = 10
ws4.column_dimensions["D"].width = 40

domain_colors = {
    "Light":       "FEF9E7",
    "Material":    "F4ECF7",
    "Transform":   "EBF5FB",
    "Spawn":       "E8F8F5",
    "Particles":   "FDF2E9",
    "Sound":       "FDEDEC",
    "Environment": "E8F6F3",
    "Post-Proc":   "F5EEF8",
    "Camera":      "EAF2F8",
}

for i, (domain, name, vtype, unity_prop) in enumerate(OUTPUTS):
    r = i + 2
    ws4.cell(row=r, column=1, value=domain)
    ws4.cell(row=r, column=2, value=name)
    ws4.cell(row=r, column=3, value=vtype)
    ws4.cell(row=r, column=4, value=unity_prop)
    for c in range(1, 5):
        cell = ws4.cell(row=r, column=c)
        style_body_cell(cell, left_align if c in (2, 4) else body_align)
        cell.fill = PatternFill("solid", fgColor=domain_colors.get(domain, WHITE))

ws4.freeze_panes = "A2"


# ── SHEET 5: PRESETS ────────────────────────────────────────────────

ws5 = wb.create_sheet("Presets")
ws5.sheet_properties.tabColor = "F39C12"

preset_headers = ["Preset Name", "Mode", "Signal", "Relationship", "Curve", "Attack", "Release", "Duration", "Description"]
for c, h in enumerate(preset_headers, 1):
    ws5.cell(row=1, column=c, value=h)
style_header(ws5, 1, len(preset_headers))

ws5.column_dimensions["A"].width = 22
ws5.column_dimensions["B"].width = 18
ws5.column_dimensions["C"].width = 8
ws5.column_dimensions["D"].width = 14
ws5.column_dimensions["E"].width = 14
ws5.column_dimensions["F"].width = 10
ws5.column_dimensions["G"].width = 10
ws5.column_dimensions["H"].width = 10
ws5.column_dimensions["I"].width = 45

# Derive mode name from signal + relationship
def _mode_name(sig, rel):
    if sig == "Cont" and rel == "Bound":
        return "Tracking"
    elif sig == "Cont" and rel == "Unbound":
        return "Threshold Trigger"
    elif sig == "Bin" and rel == "Bound":
        return "Switch"
    else:
        return "One-Shot"

for i, (name, sig, rel, curve, attack, release, dur, desc) in enumerate(PRESETS):
    r = i + 2
    mode = _mode_name(sig, rel)
    ws5.cell(row=r, column=1, value=name)
    ws5.cell(row=r, column=2, value=mode)
    ws5.cell(row=r, column=3, value=sig)
    ws5.cell(row=r, column=4, value=rel)
    ws5.cell(row=r, column=5, value=curve)
    ws5.cell(row=r, column=6, value=attack)
    ws5.cell(row=r, column=7, value=release)
    ws5.cell(row=r, column=8, value=dur)
    ws5.cell(row=r, column=9, value=desc)
    for c in range(1, 10):
        cell = ws5.cell(row=r, column=c)
        style_body_cell(cell, left_align if c in (1, 9) else body_align)
        # Tint by mode
        if rel == "Bound":
            if sig == "Cont":
                cell.fill = PatternFill("solid", fgColor=PATH_A_BG)   # Tracking
            else:
                cell.fill = PatternFill("solid", fgColor=PATH_C_BG)   # Switch
        else:
            if sig == "Cont":
                cell.fill = PatternFill("solid", fgColor=PATH_B_BG)   # Threshold Trigger
            else:
                cell.fill = PatternFill("solid", fgColor=PATH_D_BG)   # One-Shot

ws5.freeze_panes = "A2"


# ═════════════════════════════════════════════════════════════════════
#  SAVE
# ═════════════════════════════════════════════════════════════════════

output_path = r"D:\Projects Unity\MD4_2026_Mavrodiev_Develop\Docs\Interactive Design 4\Interaction Model\Interaction-Atlas.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"Atoms: {len(EXAMPLE_ATOMS)} examples + {NUM_DATA_ROWS - len(EXAMPLE_ATOMS)} empty rows")
