"""
Build the Interaction Atom Designer — an interactive calculator
where you configure one interaction atom and see:
  1. The curve shape (chart, updates live)
  2. The temporal response (chart, shows input vs output over time)
  3. Auto-generated description + coupling feel

Sheets:
  Designer   — config panel + charts + result card
  CurveData  — all curve shapes + selected curve (feeds Chart 1)
  Timeline   — input scenarios + smoothed output (feeds Chart 2)
  Lookup     — speed name → numeric factor
"""

import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════

NUM_CURVE_PTS  = 51   # X from 0 to 1, step 0.02
NUM_TIME_PTS   = 101  # T from 0 to 5, step 0.05

CURVE_NAMES = [
    "Linear", "Ease-in", "Ease-out", "Ease-in-out",
    "Inverse", "S-curve",
    "Burst", "Swell", "Snap+tail", "Overshoot", "Spike",
]

SCENARIO_NAMES = ["Approach", "Enter/Exit", "Pulse", "Oscillation"]

INPUT_NAMES = [
    "Proximity", "Gaze Direction", "Gaze Duration", "Stillness",
    "Velocity", "Slider", "Time", "Random",
    "Contact", "Button", "Toggle", "Grab", "Throw",
]
CONTINUOUS_INPUTS = [
    "Proximity", "Gaze Direction", "Gaze Duration", "Stillness",
    "Velocity", "Slider", "Time", "Random",
]

OUTPUT_NAMES = [
    "Emission Intensity", "Emission Color",
    "Point/Spot Intensity", "Light Color",
    "Base Color", "Smoothness", "Metallic", "Alpha / Opacity",
    "Dissolve", "Fresnel / Rim Glow", "Normal Map Strength",
    "Scale (uniform)", "Position", "Rotation",
    "Object Activation", "Spawn Object", "Destroy Object",
    "Emission Rate", "Particle Size", "Particle Speed", "Burst (VFX)",
    "Volume", "Pitch", "Spatial Blend", "Low-Pass Filter", "Play / Stop",
    "Fog Density", "Fog Color", "Ambient Intensity", "Ambient Color",
    "Bloom Intensity", "Color Temperature", "Saturation", "Vignette",
    "Depth of Field", "Chromatic Aberration",
    "Field of View", "Camera Shake",
]

DOMAIN_NAMES = [
    "Light", "Material", "Transform", "Spawn",
    "Particles", "Sound", "Environment", "Post-Proc", "Camera",
]

SPEED_NAMES  = ["Instant", "Fast", "Medium", "Slow", "Very Slow", "—"]
SPEED_VALUES = [1.0,       0.5,    0.2,      0.08,   0.03,         1.0]

# ── Palette ──────────────────────────────────────────────────

HEADER_BG  = "2C3E50"
HEADER_FG  = "FFFFFF"
ACCENT     = "2980B9"
ACCENT2    = "E74C3C"
LABEL_BG   = "ECF0F1"
KILLED_BG  = "E0E0E0"
KILLED_FG  = "AAAAAA"
SECTION_BG = "34495E"
RESULT_BG  = "EBF5FB"
EXPORT_BG  = "FEF9E7"

PATH_COLORS = {"Tracking": "DCEEFB", "Threshold Trigger": "D5F5E3", "Switch": "FEF5E7", "One-Shot": "FADBD8"}

header_font  = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
header_fill  = PatternFill("solid", fgColor=HEADER_BG)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

label_font   = Font(name="Calibri", bold=True, size=10, color="2C3E50")
label_fill   = PatternFill("solid", fgColor=LABEL_BG)
label_align  = Alignment(horizontal="right", vertical="center")

value_font   = Font(name="Calibri", size=11)
value_align  = Alignment(horizontal="center", vertical="center")

note_font    = Font(name="Calibri", size=9, italic=True, color="7F8C8D")

killed_fill  = PatternFill("solid", fgColor=KILLED_BG)
killed_font  = Font(name="Calibri", size=10, color=KILLED_FG)

thin_border  = Border(
    left=Side("thin", color="BDC3C7"),
    right=Side("thin", color="BDC3C7"),
    top=Side("thin", color="BDC3C7"),
    bottom=Side("thin", color="BDC3C7"),
)

section_font = Font(name="Calibri", bold=True, size=10, color=HEADER_FG)
section_fill = PatternFill("solid", fgColor=SECTION_BG)


def style_label(cell):
    cell.font = label_font
    cell.fill = label_fill
    cell.alignment = label_align
    cell.border = thin_border

def style_value(cell):
    cell.font = value_font
    cell.alignment = value_align
    cell.border = thin_border

def style_note(cell, text=None):
    if text is not None:
        cell.value = text
    cell.font = note_font
    cell.alignment = Alignment(horizontal="left", vertical="center")


# ═══════════════════════════════════════════════════════════════
#  CURVE MATH (for computing static values in CurveData)
# ═══════════════════════════════════════════════════════════════

def curve_value(name, x):
    """Compute curve value for a given curve name and x in [0,1]."""
    x = max(0.0, min(1.0, x))
    if name == "Linear":
        return x
    elif name == "Ease-in":
        return x * x
    elif name == "Ease-out":
        return 1 - (1 - x) ** 2
    elif name == "Ease-in-out":
        return 2 * x * x if x < 0.5 else 1 - (-2 * x + 2) ** 2 / 2
    elif name == "Inverse":
        return 1 - x
    elif name == "S-curve":
        return 3 * x * x - 2 * x * x * x
    elif name == "Burst":
        if x <= 0:
            return 0
        return min(1.5, (x / 0.1) * math.exp(1 - x / 0.1))
    elif name == "Swell":
        return math.sin(math.pi * x)
    elif name == "Snap+tail":
        return math.exp(-4 * x)
    elif name == "Overshoot":
        if x <= 0:
            return 0
        return min(1.5, 1.4 * (x / 0.15) * math.exp(1 - x / 0.15))
    elif name == "Spike":
        if x < 0.05:
            return x / 0.05 if 0.05 > 0 else 0
        return max(0, (1 - x) / 0.95)
    return x


# ═══════════════════════════════════════════════════════════════
#  FORMULA BUILDERS
# ═══════════════════════════════════════════════════════════════

def make_signal_formula():
    """Signal auto-fills from Input (C3)."""
    checks = ",".join([f'$C$3="{inp}"' for inp in CONTINUOUS_INPUTS])
    return f'=IF($C$3="","",IF(OR({checks}),"Continuous","Binary"))'


def make_path_formula():
    """Mode auto-fills from Signal (C4) × Relationship (C5)."""
    return (
        '=IF(OR($C$4="",$C$5=""),"",IF(AND($C$4="Continuous",$C$5="Bound"),"Tracking",'
        'IF(AND($C$4="Continuous",$C$5="Unbound"),"Threshold Trigger",'
        'IF(AND($C$4="Binary",$C$5="Bound"),"Switch",'
        'IF(AND($C$4="Binary",$C$5="Unbound"),"One-Shot","")))))'
    )


def make_path_desc_formula():
    """Human-readable mode description."""
    return (
        '=IF($C$6="Tracking","output tracks input continuously",'
        'IF($C$6="Threshold Trigger","threshold crossing triggers playback",'
        'IF($C$6="Switch","output switches between ON/OFF states",'
        'IF($C$6="One-Shot","event triggers independent playback",""))))'
    )


# ═══════════════════════════════════════════════════════════════
#  BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()


# ───────────────────────────────────────────────────────────────
#  SHEET: Lookup
# ───────────────────────────────────────────────────────────────

ws_lk = wb.create_sheet("Lookup")
ws_lk.sheet_properties.tabColor = "95A5A6"

# Speed factors
ws_lk.cell(1, 1, "Speed")
ws_lk.cell(1, 2, "Factor")
ws_lk.cell(1, 1).font = Font(bold=True)
ws_lk.cell(1, 2).font = Font(bold=True)

for i, (name, val) in enumerate(zip(SPEED_NAMES, SPEED_VALUES)):
    ws_lk.cell(i + 2, 1, name)
    ws_lk.cell(i + 2, 2, val)

ws_lk.column_dimensions["A"].width = 14
ws_lk.column_dimensions["B"].width = 10


# ───────────────────────────────────────────────────────────────
#  SHEET: CurveData
# ───────────────────────────────────────────────────────────────

ws_cd = wb.create_sheet("CurveData")
ws_cd.sheet_properties.tabColor = "2980B9"

# Headers
ws_cd.cell(1, 1, "X")
ws_cd.cell(1, 1).font = Font(bold=True)
for ci, cname in enumerate(CURVE_NAMES):
    col = ci + 2  # B=2, C=3, ...
    ws_cd.cell(1, col, cname)
    ws_cd.cell(1, col).font = Font(bold=True, size=9)
    ws_cd.column_dimensions[get_column_letter(col)].width = 12

sel_col = len(CURVE_NAMES) + 2  # M = 13
ws_cd.cell(1, sel_col, "▶ Selected")
ws_cd.cell(1, sel_col).font = Font(bold=True, color=ACCENT)
ws_cd.column_dimensions[get_column_letter(sel_col)].width = 12
ws_cd.column_dimensions["A"].width = 6

# Data rows: static values for each curve, formula for Selected
for ri in range(NUM_CURVE_PTS):
    r = ri + 2  # data starts at row 2
    x = ri / (NUM_CURVE_PTS - 1)  # 0 to 1

    ws_cd.cell(r, 1, round(x, 4))

    for ci, cname in enumerate(CURVE_NAMES):
        col = ci + 2
        ws_cd.cell(r, col, round(curve_value(cname, x), 6))

    # Selected: INDEX/MATCH from Designer!$C$8
    first_col = get_column_letter(2)         # B
    last_col  = get_column_letter(sel_col - 1)  # L (last curve col)
    ws_cd.cell(r, sel_col,
        f'=IFERROR(INDEX(${first_col}{r}:${last_col}{r},1,'
        f'MATCH(Designer!$C$8,${first_col}$1:${last_col}$1,0)),0)'
    )

ws_cd.freeze_panes = "A2"


# ───────────────────────────────────────────────────────────────
#  SHEET: Timeline
# ───────────────────────────────────────────────────────────────

ws_tl = wb.create_sheet("Timeline")
ws_tl.sheet_properties.tabColor = "E74C3C"

# Headers
tl_headers = ["Time", "Approach", "Enter/Exit", "Pulse", "Oscillation",
              "▶ Input", "▶ Target", "▶ Output"]
for ci, h in enumerate(tl_headers):
    c = ci + 1
    cell = ws_tl.cell(1, c, h)
    cell.font = Font(bold=True, size=9, color=ACCENT if "▶" in h else "000000")
    ws_tl.column_dimensions[get_column_letter(c)].width = 12

# -- Reference cells on Designer sheet --
# C3=Input, C4=Signal, C5=Rel, C6=Mode, C8=Curve
# C9=InMin, E9=InMax, C10=OutMin, E10=OutMax
# C11=Attack, C12=Release, C13=Duration, C17=Scenario

# Curve lookup: given a normalized value ref, return the curve value
# from CurveData's Selected column (col M = sel_col)
sel_col_letter = get_column_letter(sel_col)  # "M"

def curve_lookup(val_ref):
    """INDEX into CurveData's Selected column for a given normalized [0-1] value."""
    return (
        f'IFERROR(INDEX(CurveData!${sel_col_letter}$2:${sel_col_letter}${NUM_CURVE_PTS+1},'
        f'MAX(1,MATCH(MAX(0,MIN(1,{val_ref})),'
        f'CurveData!$A$2:$A${NUM_CURVE_PTS+1},1))),0)'
    )

for ri in range(NUM_TIME_PTS):
    r = ri + 2
    t = round(ri * 5.0 / (NUM_TIME_PTS - 1), 4)

    # A: Time
    ws_tl.cell(r, 1, t)

    # B: Approach — ramps from 0 to 1 between t=0.5 and t=3.5
    ws_tl.cell(r, 2, f'=MIN(1,MAX(0,(A{r}-0.5)/3))')

    # C: Enter/Exit — step to 1 at t=1, step to 0 at t=3.5
    ws_tl.cell(r, 3, f'=IF(AND(A{r}>=1,A{r}<3.5),1,0)')

    # D: Pulse — step to 1 at t=1, step to 0 at t=1.3
    ws_tl.cell(r, 4, f'=IF(AND(A{r}>=1,A{r}<1.3),1,0)')

    # E: Oscillation — smooth sine wave
    ws_tl.cell(r, 5, f'=0.5+0.5*SIN(2*PI()*A{r}/2)')

    # F: Selected Input — pick scenario column via MATCH
    ws_tl.cell(r, 6,
        f'=IFERROR(INDEX(B{r}:E{r},1,'
        f'MATCH(Designer!$C$17,B$1:E$1,0)),0)'
    )

    # G: Target
    # For bound (Tracking/Switch): curve(input_signal)
    # For unbound (Threshold Trigger/One-Shot): if in playback [t=1..1+dur], curve(progress); else 0
    bound_val   = curve_lookup(f'F{r}')
    progress    = f'(A{r}-1)/Designer!$C$13'
    unbound_val = curve_lookup(progress)
    playback_check = f'AND(A{r}>=1,A{r}<=1+Designer!$C$13)'

    ws_tl.cell(r, 7,
        f'=IF(OR(Designer!$C$6="Tracking",Designer!$C$6="Switch"),'
        f'{bound_val},'
        f'IF({playback_check},'
        f'{unbound_val},'
        f'0))'
    )

    # H: Smoothed Output
    # For unbound: = Target (no extra smoothing, curve IS the shape)
    # For bound: exponential smoothing with attack/release
    if ri == 0:
        ws_tl.cell(r, 8, f'=G{r}')  # initialize
    else:
        atk  = f'IFERROR(VLOOKUP(Designer!$C$11,Lookup!$A$2:$B$7,2,FALSE),1)'
        rel  = f'IFERROR(VLOOKUP(Designer!$C$12,Lookup!$A$2:$B$7,2,FALSE),1)'
        prev = f'H{r-1}'
        tgt  = f'G{r}'
        ws_tl.cell(r, 8,
            f'=IF(OR(Designer!$C$6="Threshold Trigger",Designer!$C$6="One-Shot"),'
            f'{tgt},'
            f'{prev}+({tgt}-{prev})*IF({tgt}>{prev},{atk},{rel}))'
        )

ws_tl.freeze_panes = "A2"


# ───────────────────────────────────────────────────────────────
#  SHEET: Designer (main — must be first/active)
# ───────────────────────────────────────────────────────────────

ws = wb.active
ws.title = "Designer"
ws.sheet_properties.tabColor = "8E44AD"

# Column widths
col_widths = {"A": 16, "B": 4, "C": 18, "D": 14, "E": 14, "F": 14,
              "G": 2}  # spacer before charts
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# ── Title ──
ws.merge_cells("A1:F1")
title_cell = ws.cell(1, 1, "INTERACTION ATOM DESIGNER")
title_cell.font = Font(name="Calibri", bold=True, size=14, color=HEADER_FG)
title_cell.fill = header_fill
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# ── Config panel ──

def config_row(row, label, value_cell_col=3, value=None, note_col=None, note=None, row_height=22):
    """Write a label + value cell for the config panel."""
    ws.row_dimensions[row].height = row_height
    c_label = ws.cell(row, 1, label)
    style_label(c_label)
    # Fill label cells B
    ws.cell(row, 2).fill = label_fill
    ws.cell(row, 2).border = thin_border

    if value is not None:
        c_val = ws.cell(row, value_cell_col, value)
        style_value(c_val)

    if note_col and note:
        style_note(ws.cell(row, note_col, note))

def section_row(row):
    ws.row_dimensions[row].height = 6
    for c in range(1, 7):
        cell = ws.cell(row, c)
        cell.fill = section_fill
        cell.border = thin_border

# Row 2: spacer
ws.row_dimensions[2].height = 6

# Row 3: INPUT
config_row(3, "INPUT", value="Proximity")
style_note(ws.cell(3, 4, "← select input source"))

# Row 4: SIGNAL (auto)
config_row(4, "SIGNAL")
ws.cell(4, 3, make_signal_formula())
style_value(ws.cell(4, 3))
ws.cell(4, 3).font = Font(name="Calibri", size=11, italic=True, color="7F8C8D")
style_note(ws.cell(4, 4, "auto from Input"))

# Row 5: RELATIONSHIP
config_row(5, "RELATIONSHIP", value="Bound")

# Row 6: MODE (auto)
config_row(6, "MODE")
ws.cell(6, 3, make_path_formula())
style_value(ws.cell(6, 3))
ws.cell(6, 3).font = Font(name="Calibri", size=11, bold=True, color=ACCENT)
ws.cell(6, 4, make_path_desc_formula())
ws.cell(6, 4).font = note_font

# Row 7: section separator
section_row(7)

# Row 8: CURVE
config_row(8, "CURVE", value="Linear")

# Row 9: INPUT RANGE
ws.row_dimensions[9].height = 22
ws.cell(9, 1, "INPUT RANGE").font = label_font
ws.cell(9, 1).fill = label_fill
ws.cell(9, 1).alignment = label_align
ws.cell(9, 1).border = thin_border
ws.cell(9, 2, "min").font = note_font
ws.cell(9, 2).alignment = value_align
ws.cell(9, 2).fill = label_fill
ws.cell(9, 2).border = thin_border
ws.cell(9, 3, 0)
style_value(ws.cell(9, 3))
# D9: dynamic label "max" or "threshold"
ws.cell(9, 4, '=IF($C$6="Threshold Trigger","threshold","max")')
ws.cell(9, 4).font = note_font
ws.cell(9, 4).alignment = value_align
ws.cell(9, 4).border = thin_border
ws.cell(9, 5, 5)
style_value(ws.cell(9, 5))

# Row 10: OUTPUT RANGE
ws.row_dimensions[10].height = 22
ws.cell(10, 1, "OUTPUT RANGE").font = label_font
ws.cell(10, 1).fill = label_fill
ws.cell(10, 1).alignment = label_align
ws.cell(10, 1).border = thin_border
ws.cell(10, 2, "min").font = note_font
ws.cell(10, 2).alignment = value_align
ws.cell(10, 2).fill = label_fill
ws.cell(10, 2).border = thin_border
ws.cell(10, 3, 0)
style_value(ws.cell(10, 3))
ws.cell(10, 4, "max").font = note_font
ws.cell(10, 4).alignment = value_align
ws.cell(10, 4).border = thin_border
ws.cell(10, 5, 1)
style_value(ws.cell(10, 5))

# Row 11: ATTACK
config_row(11, "ATTACK", value="—")

# Row 12: RELEASE
config_row(12, "RELEASE", value="—")

# Row 13: DURATION
config_row(13, "DURATION", value="—")
ws.cell(13, 4, '=IF(OR($C$6="Tracking",$C$6="Switch"),"— continuous","sec")')
ws.cell(13, 4).font = note_font
ws.cell(13, 4).alignment = Alignment(horizontal="left", vertical="center")

# Row 14: section separator
section_row(14)

# Row 15: OUTPUT TARGET
config_row(15, "OUTPUT", value="Emission Intensity")

# Row 16: DOMAIN
config_row(16, "DOMAIN", value="Light")

# Row 17: SCENARIO
config_row(17, "SCENARIO", value="Approach")
style_note(ws.cell(17, 4, "for timeline chart →"))

# Row 18: section separator
section_row(18)

# ── Result Card ──

ws.merge_cells("A19:F19")
ws.cell(19, 1, "RESULT").font = Font(bold=True, size=11, color=HEADER_FG)
ws.cell(19, 1).fill = PatternFill("solid", fgColor=ACCENT)
ws.cell(19, 1).alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[19].height = 24

# Row 20: description line
ws.merge_cells("A20:F20")
ws.cell(20, 1,
    '=IF($C$6="Tracking",'
    'CONCATENATE("When ",$C$3," changes, ",$C$15," tracks it with a ",$C$8," curve."),'
    'IF($C$6="Threshold Trigger",'
    'CONCATENATE("When ",$C$3," crosses ",$E$9,", ",$C$15," plays a ",$C$8," shape for ",$C$13,"."),'
    'IF($C$6="Switch",'
    'CONCATENATE("When ",$C$3," activates, ",$C$15," transitions ",$C$11," on / ",$C$12," off."),'
    'CONCATENATE("When ",$C$3," fires, ",$C$15," plays a ",$C$8," shape for ",$C$13,"."))))'
)
ws.cell(20, 1).font = Font(name="Calibri", size=10)
ws.cell(20, 1).fill = PatternFill("solid", fgColor=RESULT_BG)
ws.cell(20, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[20].height = 36

# Row 21: parameter summary
ws.merge_cells("A21:F21")
ws.cell(21, 1,
    '=CONCATENATE('
    '"Mode: ",C6,'
    'IF(C8<>"—",CONCATENATE("  ·  Curve: ",C8),""),'
    'IF(AND(C11<>"—",C11<>""),CONCATENATE("  ·  Attack: ",C11),""),'
    'IF(AND(C12<>"—",C12<>""),CONCATENATE("  ·  Release: ",C12),""),'
    'IF(AND(C13<>"—",C13<>""),CONCATENATE("  ·  Duration: ",C13),""))'
)
ws.cell(21, 1).font = Font(name="Calibri", size=9, color="7F8C8D")
ws.cell(21, 1).fill = PatternFill("solid", fgColor=RESULT_BG)
ws.cell(21, 1).alignment = Alignment(horizontal="left", vertical="center")

# Row 22: coupling quality suggestion
ws.merge_cells("A22:F22")
ws.cell(22, 1,
    '=CONCATENATE("Coupling: ",'
    'IF($C$6="Tracking",'
    '  CONCATENATE('
    '    IF($C$8="Linear","Attentive",IF($C$8="Inverse","Reluctant",'
    '      IF($C$8="Ease-out","Responsive",IF($C$8="Ease-in","Delayed","Shaped")))),'
    '    ", ",'
    '    IF(OR($C$11="—",$C$11="Instant",$C$11=""),'
    '      IF(OR($C$12="—",$C$12="Instant",$C$12=""),"crisp","lingering"),'
    '      IF(OR($C$12="—",$C$12="Instant",$C$12=""),"reluctant onset, crisp","atmospheric"))),'
    'IF($C$6="Switch",'
    '  CONCATENATE('
    '    IF(OR($C$11="Instant",$C$11="Fast"),"Eager","Reluctant"),'
    '    ", ",'
    '    IF(OR($C$12="Instant",$C$12="Fast"),"crisp","lingering")),'
    'IF($C$6="One-Shot",'
    '  IF($C$8="Burst","Explosive, eager",'
    '    IF($C$8="Swell","Organic, breathing",'
    '      IF($C$8="Snap+tail","Sharp, then lingering",'
    '        IF($C$8="Overshoot","Playful, bouncy","Expressive")))),'
    '"Threshold-triggered"))))'
)
ws.cell(22, 1).font = Font(name="Calibri", bold=True, size=10, color=ACCENT)
ws.cell(22, 1).fill = PatternFill("solid", fgColor=RESULT_BG)
ws.cell(22, 1).alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[22].height = 24

# ── Conditional Formatting: gray out killed fields ──

def kill_cells(cell_range, path_formula):
    """Apply killed style when path formula is true."""
    rule = FormulaRule(formula=[path_formula], fill=killed_fill, font=killed_font)
    ws.conditional_formatting.add(cell_range, rule)

# Curve (row 8): killed when Mode = Switch (binary+bound — no mapping curve)
kill_cells("A8:E8", '$C$6="Switch"')

# Input Range Min (C9): killed when Mode ≠ Tracking
kill_cells("A9:C9", 'OR($C$6="Threshold Trigger",$C$6="Switch",$C$6="One-Shot")')

# Input Range Max/Threshold (D9:E9): killed when Mode = Switch or One-Shot
kill_cells("D9:E9", 'OR($C$6="Switch",$C$6="One-Shot")')

# Attack (row 11): killed when Mode = Threshold Trigger or One-Shot (unbound — baked in curve)
kill_cells("A11:E11", 'OR($C$6="Threshold Trigger",$C$6="One-Shot")')

# Release (row 12): killed when Mode = Threshold Trigger or One-Shot
kill_cells("A12:E12", 'OR($C$6="Threshold Trigger",$C$6="One-Shot")')

# Duration (row 13): killed when Mode = Tracking or Switch (continuous — no fixed duration)
kill_cells("A13:E13", 'OR($C$6="Tracking",$C$6="Switch")')

# ── Mode color tinting on Mode cell ──
for path_letter, color in PATH_COLORS.items():
    rule = FormulaRule(
        formula=[f'$C$6="{path_letter}"'],
        fill=PatternFill("solid", fgColor=color),
    )
    ws.conditional_formatting.add("C6", rule)


# ── Data Validations (dropdowns) ──

def add_dropdown(cell_range, options, prompt_title="", prompt_text=""):
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(options) + '"',
        allow_blank=True,
    )
    if prompt_title:
        dv.promptTitle = prompt_title
        dv.prompt = prompt_text
    ws.add_data_validation(dv)
    dv.add(cell_range)

add_dropdown("C3",  INPUT_NAMES,    "Input",        "What triggers this interaction?")
add_dropdown("C5",  ["Bound", "Unbound"], "Relationship",
    "Bound = output follows input. Unbound = output plays independently.")
add_dropdown("C8",  CURVE_NAMES,    "Curve",
    "Mapping shape. Bound: input→output. Unbound: time→output response.")
add_dropdown("C11", SPEED_NAMES,    "Attack",       "How fast does output reach target?")
add_dropdown("C12", SPEED_NAMES,    "Release",      "How fast does output return to rest?")
add_dropdown("C17", SCENARIO_NAMES, "Scenario",     "Input pattern for timeline chart.")

# Output dropdown — long list, might hit 255 char limit
# Split into shorter list or use reference
output_str = ",".join(OUTPUT_NAMES)
if len(output_str) <= 250:
    add_dropdown("C15", OUTPUT_NAMES, "Output", "What property changes?")
else:
    # Reference a range instead — put output names on Lookup sheet
    out_start = 10
    ws_lk.cell(out_start, 1, "Outputs")
    ws_lk.cell(out_start, 1).font = Font(bold=True)
    for oi, oname in enumerate(OUTPUT_NAMES):
        ws_lk.cell(out_start + 1 + oi, 1, oname)
    out_end = out_start + len(OUTPUT_NAMES)
    dv_out = DataValidation(
        type="list",
        formula1=f"Lookup!$A${out_start+1}:$A${out_end}",
        allow_blank=True,
    )
    dv_out.promptTitle = "Output"
    dv_out.prompt = "What property changes?"
    ws.add_data_validation(dv_out)
    dv_out.add("C15")

add_dropdown("C16", DOMAIN_NAMES, "Domain", "Perceptual category.")


# ── Export Row ──

ws.row_dimensions[24].height = 20
ws.merge_cells("A24:F24")
ws.cell(24, 1, "COPY TO ATLAS →  select row 26, copy, paste into Atlas sheet")
ws.cell(24, 1).font = Font(name="Calibri", size=9, italic=True, color="7F8C8D")
ws.cell(24, 1).fill = PatternFill("solid", fgColor=EXPORT_BG)
ws.cell(24, 1).alignment = Alignment(horizontal="left", vertical="center")

# Row 25: Atlas column headers
atlas_headers = ["#", "Name", "Input", "Signal", "Rel", "Mode",
                 "Curve", "In Min", "In Max", "Out Min", "Out Max",
                 "Attack", "Release", "Duration", "Output", "Domain", "Feel"]
for ci, h in enumerate(atlas_headers):
    c = ci + 1
    cell = ws.cell(25, c, h)
    cell.font = Font(name="Calibri", bold=True, size=8, color=HEADER_FG)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Row 26: formulas referencing config
atlas_formulas = [
    "",                                             # #
    '=CONCATENATE($C$3," → ",$C$15)',               # Name
    '=$C$3',                                        # Input
    '=$C$4',                                        # Signal
    '=$C$5',                                        # Relationship
    '=$C$6',                                        # Mode
    '=$C$8',                                        # Curve
    '=IF(OR($C$6="Threshold Trigger",$C$6="Switch",$C$6="One-Shot"),"—",$C$9)',  # In Min
    '=IF(OR($C$6="Switch",$C$6="One-Shot"),"—",$E$9)',          # In Max
    '=$C$10',                                       # Out Min
    '=$E$10',                                       # Out Max
    '=IF(OR($C$6="Threshold Trigger",$C$6="One-Shot"),"—",$C$11)',          # Attack
    '=IF(OR($C$6="Threshold Trigger",$C$6="One-Shot"),"—",$C$12)',          # Release
    '=IF(OR($C$6="Tracking",$C$6="Switch"),"—",$C$13)',          # Duration
    '=$C$15',                                       # Output
    '=$C$16',                                       # Domain
    '',                                             # Feel (manual)
]
for ci, formula in enumerate(atlas_formulas):
    c = ci + 1
    cell = ws.cell(26, c, formula if formula else "")
    cell.font = Font(name="Calibri", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor=EXPORT_BG)
    cell.border = thin_border


# ───────────────────────────────────────────────────────────────
#  CHARTS
# ───────────────────────────────────────────────────────────────

# ── Chart 1: Curve Shape ──

chart1 = ScatterChart()
chart1.title = "Curve Shape"
chart1.x_axis.title = "Normalized Input (bound)  /  Playback Progress (unbound)"
chart1.y_axis.title = "Output (normalized)"
chart1.style = 2
chart1.width = 18
chart1.height = 11
chart1.x_axis.scaling.min = 0
chart1.x_axis.scaling.max = 1
chart1.y_axis.scaling.min = -0.05
chart1.y_axis.scaling.max = 1.5
chart1.x_axis.numFmt = '0.0'
chart1.y_axis.numFmt = '0.00'
chart1.legend = None

x_ref = Reference(ws_cd, min_col=1, min_row=2, max_row=NUM_CURVE_PTS + 1)
y_ref = Reference(ws_cd, min_col=sel_col, min_row=2, max_row=NUM_CURVE_PTS + 1)

s_curve = Series(y_ref, x_ref, title="Selected Curve")
s_curve.smooth = True
s_curve.graphicalProperties.line.solidFill = ACCENT
s_curve.graphicalProperties.line.width = 28000  # ~2.2pt
s_curve.marker = Marker(symbol='none')
chart1.series.append(s_curve)

# Add a subtle reference line at y=1 (using a 2-point series)
# Write helper data for the reference line
ref_line_col = sel_col + 2  # column after a gap
ws_cd.cell(1, ref_line_col, "ref_x")
ws_cd.cell(1, ref_line_col + 1, "ref_y")
ws_cd.cell(2, ref_line_col, 0)
ws_cd.cell(2, ref_line_col + 1, 1)
ws_cd.cell(3, ref_line_col, 1)
ws_cd.cell(3, ref_line_col + 1, 1)

ref_x = Reference(ws_cd, min_col=ref_line_col, min_row=2, max_row=3)
ref_y = Reference(ws_cd, min_col=ref_line_col + 1, min_row=2, max_row=3)
s_ref = Series(ref_y, ref_x, title="y=1")
s_ref.graphicalProperties.line.solidFill = "BDC3C7"
s_ref.graphicalProperties.line.width = 10000
s_ref.graphicalProperties.line.dashStyle = "dash"
s_ref.marker = Marker(symbol='none')
chart1.series.append(s_ref)

ws.add_chart(chart1, "H2")

# ── Chart 2: Timeline Response ──

chart2 = ScatterChart()
chart2.title = "Timeline Response"
chart2.x_axis.title = "Time (seconds)"
chart2.y_axis.title = "Value (normalized 0–1)"
chart2.style = 2
chart2.width = 18
chart2.height = 11
chart2.x_axis.scaling.min = 0
chart2.x_axis.scaling.max = 5
chart2.y_axis.scaling.min = -0.05
chart2.y_axis.scaling.max = 1.5
chart2.x_axis.numFmt = '0.0'
chart2.y_axis.numFmt = '0.00'

time_ref = Reference(ws_tl, min_col=1, min_row=2, max_row=NUM_TIME_PTS + 1)

# Input signal (dashed gray)
input_ref = Reference(ws_tl, min_col=6, min_row=2, max_row=NUM_TIME_PTS + 1)
s_input = Series(input_ref, time_ref, title="Input")
s_input.graphicalProperties.line.solidFill = "BDC3C7"
s_input.graphicalProperties.line.width = 15000
s_input.graphicalProperties.line.dashStyle = "dash"
s_input.marker = Marker(symbol='none')
chart2.series.append(s_input)

# Target (thin, dotted, subtle — shows pre-smoothing)
target_ref = Reference(ws_tl, min_col=7, min_row=2, max_row=NUM_TIME_PTS + 1)
s_target = Series(target_ref, time_ref, title="Target")
s_target.graphicalProperties.line.solidFill = "F5B7B1"
s_target.graphicalProperties.line.width = 10000
s_target.graphicalProperties.line.dashStyle = "dot"
s_target.marker = Marker(symbol='none')
chart2.series.append(s_target)

# Smoothed output (bold red — the actual output)
output_ref = Reference(ws_tl, min_col=8, min_row=2, max_row=NUM_TIME_PTS + 1)
s_output = Series(output_ref, time_ref, title="Output")
s_output.graphicalProperties.line.solidFill = ACCENT2
s_output.graphicalProperties.line.width = 28000
s_output.marker = Marker(symbol='none')
chart2.series.append(s_output)

ws.add_chart(chart2, "H17")


# ═══════════════════════════════════════════════════════════════
#  SAVE
# ═══════════════════════════════════════════════════════════════

output_path = (
    r"D:\Projects Unity\MD4_2026_Mavrodiev_Develop"
    r"\Docs\Interactive Design 4\Interaction Model"
    r"\Interaction-Atom-Designer.xlsx"
)
wb.save(output_path)

print(f"Saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"CurveData: {NUM_CURVE_PTS} points × {len(CURVE_NAMES)} curves + selected")
print(f"Timeline:  {NUM_TIME_PTS} points × {len(SCENARIO_NAMES)} scenarios")
print("Open in Excel — change any dropdown and charts update live.")
